import json
from datetime import UTC, date, datetime, timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.security import create_access_token
from app.models import (
    AuditLog,
    DisciplineCase,
    EvaluationAppeal,
    EvaluationCriterion,
    EvaluationCycle,
    EvaluationEvidence,
    EvaluationScoreEvent,
    Member,
    MemberCycleRole,
    MemberEvaluation,
    MemberEvaluationBreakdown,
    User,
)
from app.services.evaluation_report import EvaluationReportService


def _auth_header(user_id: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token(user_id)}"}


def _user(test_db: Session, username: str, role: str, *, email: str | None = None) -> User:
    user = User(
        id=f"user-{username}",
        username=username,
        password_hash="hashed",
        full_name=username.title(),
        role=role,
        email=email,
        is_active=True,
    )
    test_db.add(user)
    test_db.commit()
    return user


def _member(test_db: Session, mssv: str, *, ban: str = "TECH") -> Member:
    member = Member(
        id=f"member-{mssv}",
        mssv=mssv,
        name=f"Member {mssv}",
        ban=ban,
        email=f"{mssv.lower()}@example.test",
    )
    test_db.add(member)
    test_db.commit()
    return member


def _cycle(test_db: Session, *, status: str = "LOCKED") -> EvaluationCycle:
    cycle = EvaluationCycle(
        id="cycle-phase6",
        code="2026-05-P6",
        name="Phase 6 cycle",
        type="MONTHLY",
        start_date=date(2026, 5, 1),
        end_date=date(2026, 5, 31),
        status=status,
    )
    test_db.add(cycle)
    test_db.commit()
    return cycle


def _criterion(
    test_db: Session,
    code: str,
    *,
    component: str,
    requires_evidence: bool = False,
) -> EvaluationCriterion:
    criterion = EvaluationCriterion(
        id=f"criterion-{code}",
        code=code,
        name=code,
        component=component,
        unit_scope="ALL",
        max_score=30.0,
        score_method="MANUAL",
        requires_evidence=requires_evidence,
    )
    test_db.add(criterion)
    test_db.commit()
    return criterion


def _evaluation(
    test_db: Session,
    cycle: EvaluationCycle,
    member: Member,
    *,
    total: float,
    classification: str,
    status: str = "LOCKED",
    blockers: list | None = None,
    attendance_rate: float | None = 95.0,
) -> MemberEvaluation:
    row = MemberEvaluation(
        id=f"eval-{member.mssv}",
        cycle_id=cycle.id,
        member_id=member.id,
        component_i_score=25.0,
        component_ii_score=15.0,
        component_iii_a_score=20.0,
        component_iii_b_score=total - 60.0,
        total_score=total,
        preliminary_classification=classification,
        final_classification=classification,
        status=status,
        attendance_rate=attendance_rate,
        blockers_json=json.dumps(blockers or []),
        computed_at=datetime.now(UTC),
        approved_at=datetime.now(UTC),
    )
    test_db.add(row)
    test_db.commit()
    return row


def _role(
    test_db: Session,
    cycle: EvaluationCycle,
    member: Member,
    unit_code: str,
    *,
    is_primary: bool = True,
) -> MemberCycleRole:
    role = MemberCycleRole(
        cycle_id=cycle.id,
        member_id=member.id,
        unit_code=unit_code,
        role_type="MEMBER",
        role_title="Member",
        participation_weight=1.0,
        is_primary=is_primary,
    )
    test_db.add(role)
    test_db.commit()
    return role


def _seed_phase6(test_db: Session):
    cycle = _cycle(test_db)
    criterion = _criterion(test_db, "I.1", component="I", requires_evidence=True)
    tech_member = _member(test_db, "MTEC001", ban="TECH")
    hr_member = _member(test_db, "MTEC002", ban="HR")
    _role(test_db, cycle, tech_member, "TECH")
    _role(test_db, cycle, hr_member, "HR")
    tech_eval = _evaluation(
        test_db,
        cycle,
        tech_member,
        total=90.0,
        classification="EXCELLENT",
    )
    hr_eval = _evaluation(
        test_db,
        cycle,
        hr_member,
        total=55.0,
        classification="NEEDS_IMPROVEMENT",
        blockers=[{"code": "INTERNAL_WARNING"}],
        attendance_rate=70.0,
    )
    test_db.add(
        MemberEvaluationBreakdown(
            member_evaluation_id=tech_eval.id,
            cycle_id=cycle.id,
            member_id=tech_member.id,
            criterion_id=criterion.id,
            criterion_code=criterion.code,
            component=criterion.component,
            raw_score=25.0,
            final_score=25.0,
            max_score_snapshot=30.0,
            evidence_count=1,
        )
    )
    event = EvaluationScoreEvent(
        cycle_id=cycle.id,
        member_id=hr_member.id,
        criterion_id=criterion.id,
        criterion_code=criterion.code,
        component=criterion.component,
        event_type="MANUAL_SCORE",
        score_delta=10.0,
    )
    test_db.add(event)
    test_db.flush()
    test_db.add(
        EvaluationEvidence(
            cycle_id=cycle.id,
            member_id=tech_member.id,
            criterion_id=criterion.id,
            evidence_type="URL",
            title="Evidence",
            url="https://example.test/evidence",
            status="VERIFIED",
        )
    )
    test_db.add(
        DisciplineCase(
            cycle_id=cycle.id,
            member_id=hr_member.id,
            case_code="CASE-HR",
            case_type="WARNING",
            severity="MEDIUM",
            title="Warning",
            blocker_code="INTERNAL_WARNING",
        )
    )
    appeal = EvaluationAppeal(
        cycle_id=cycle.id,
        member_id=hr_member.id,
        member_evaluation_id=hr_eval.id,
        criterion_id=criterion.id,
        criterion_code=criterion.code,
        appeal_type="PROFESSIONAL_SCORE",
        content="Review please",
        status="ACCEPTED",
        created_at=datetime.now(UTC) - timedelta(hours=4),
        resolved_at=datetime.now(UTC),
    )
    test_db.add(appeal)
    test_db.commit()
    return cycle, tech_member, hr_member


def test_cycle_dashboard_counts_members(test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)

    dashboard = EvaluationReportService(test_db).get_cycle_dashboard(cycle.id)

    assert dashboard["totalMembers"] == 2
    assert dashboard["computedMembers"] == 2
    assert dashboard["approvedMembers"] == 2
    assert dashboard["lockedMembers"] == 2


def test_cycle_dashboard_classification_distribution(test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)

    dashboard = EvaluationReportService(test_db).get_cycle_dashboard(cycle.id)

    assert dashboard["classificationDistribution"]["EXCELLENT"] == 1
    assert dashboard["classificationDistribution"]["NEEDS_IMPROVEMENT"] == 1


def test_component_averages_are_correct(test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)

    averages = EvaluationReportService(test_db).get_component_averages(cycle.id)

    assert averages["componentAverages"]["I"] == 25.0
    assert averages["componentAverages"]["II"] == 15.0


def test_member_report_contains_breakdowns_and_blockers(test_db: Session):
    cycle, tech_member, hr_member = _seed_phase6(test_db)

    service = EvaluationReportService(test_db)
    tech_report = service.get_member_report(cycle.id, tech_member.id)
    hr_report = service.get_member_report(cycle.id, hr_member.id)

    assert tech_report["breakdowns"][0]["criterionCode"] == "I.1"
    assert hr_report["blockers"][0]["code"] == "INTERNAL_WARNING"


def test_unit_report_filters_members_by_unit(test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)

    report = EvaluationReportService(test_db).get_unit_report(cycle.id, "TECH")

    assert report["totalMembers"] == 1
    assert report["members"][0]["unitCode"] == "TECH"


def test_risk_report_counts_discipline_cases(test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)

    report = EvaluationReportService(test_db).get_risk_report(cycle.id)

    assert report["summary"]["disciplineCases"] == 1
    assert report["summary"]["attendanceUnder80"] == 1
    assert report["summary"]["internalWarnings"] >= 1


def test_appeal_report_counts_statuses(test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)

    report = EvaluationReportService(test_db).get_appeal_report(cycle.id)

    assert report["totalAppeals"] == 1
    assert report["acceptedAppeals"] == 1
    assert report["averageResolutionHours"] == 4.0


def test_member_cannot_view_cycle_dashboard(client: TestClient, test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)
    user = _user(test_db, "member_user", "member")

    response = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/dashboard",
        headers=_auth_header(user.id),
    )

    assert response.status_code == 403


def test_bvh_can_view_cycle_dashboard(client: TestClient, test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)
    user = _user(test_db, "bvh", "bvh_discipline")

    response = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/dashboard",
        headers=_auth_header(user.id),
    )

    assert response.status_code == 200
    assert response.json()["data"]["totalMembers"] == 2


def test_bcm_can_view_unit_report_only(client: TestClient, test_db: Session):
    cycle, tech_member, _ = _seed_phase6(test_db)
    user = _user(test_db, tech_member.mssv, "bcm")

    allowed = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/units/TECH",
        headers=_auth_header(user.id),
    )
    forbidden = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/units/HR",
        headers=_auth_header(user.id),
    )

    assert allowed.status_code == 200
    assert forbidden.status_code == 403


def test_export_members_csv_has_expected_columns(client: TestClient, test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)
    user = _user(test_db, "exporter", "bvh_hr")

    response = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/exports/members.csv",
        headers=_auth_header(user.id),
    )

    assert response.status_code == 200
    first_line = response.content.decode("utf-8-sig").splitlines()[0]
    assert "cycle_id,member_evaluation_id,mssv,name,unit_code" in first_line
    assert test_db.scalar(select(AuditLog).where(AuditLog.action == "EXPORT_EVALUATION_MEMBERS_CSV"))


def test_export_members_csv_respects_filters(client: TestClient, test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)
    user = _user(test_db, "exporter2", "bvh_hr")

    response = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/exports/members.csv?unitCode=TECH",
        headers=_auth_header(user.id),
    )

    csv_text = response.content.decode("utf-8-sig")
    assert "MTEC001" in csv_text
    assert "MTEC002" not in csv_text


def test_export_members_xlsx_has_expected_sheets(client: TestClient, test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)
    user = _user(test_db, "xlsx", "bcn")

    response = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/exports/members.xlsx",
        headers=_auth_header(user.id),
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert {"Overview", "Members", "Breakdowns", "Blockers", "Appeals"}.issubset(
        set(workbook.sheetnames)
    )


def test_member_can_export_own_report(client: TestClient, test_db: Session):
    cycle, tech_member, _ = _seed_phase6(test_db)
    user = _user(test_db, tech_member.mssv, "member")

    response = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/members/{tech_member.id}/exports/report.docx",
        headers=_auth_header(user.id),
    )

    assert response.status_code == 200
    assert response.content.startswith(b"PK")


def test_member_cannot_export_other_member_report(client: TestClient, test_db: Session):
    cycle, tech_member, hr_member = _seed_phase6(test_db)
    user = _user(test_db, tech_member.mssv, "member")

    response = client.get(
        f"/api/v2/evaluations/reports/cycles/{cycle.id}/members/{hr_member.id}/exports/report.docx",
        headers=_auth_header(user.id),
    )

    assert response.status_code == 403


def test_locked_cycle_report_is_cacheable(test_db: Session):
    cycle, _, _ = _seed_phase6(test_db)
    service = EvaluationReportService(test_db)

    first = service.get_cycle_dashboard(cycle.id)
    second = service.get_cycle_dashboard(cycle.id)

    assert first["cache"]["cacheable"] is True
    assert second["cache"]["cacheable"] is True
    assert second["cache"]["cachedAt"]
